"""
STOPLIGHT REGISTRATION SYSTEM
Director-Grade Enterprise Application — Render Ready
"""

import os
import io
import json
import uuid
import base64
import secrets
import string
from datetime import datetime, date, timedelta
from functools import wraps

from flask import (Flask, render_template, request, redirect, url_for,
                   flash, session, jsonify, send_file, abort)
from flask_sqlalchemy import SQLAlchemy
from flask_login import (LoginManager, UserMixin, login_user, logout_user,
                          login_required, current_user)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from sqlalchemy import func, or_

import qrcode
import openpyxl

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                  Spacer, Image as RLImage, HRFlowable)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# ─────────────────────────────────────────────
# APP CONFIGURATION
# ─────────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.config['SQLALCHEMY_DATABASE_URI'] = (
    f"sqlite:///{os.path.join(BASE_DIR, 'stoplight.db')}"
)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = os.path.join(BASE_DIR, 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'


# ─────────────────────────────────────────────
# DATABASE MODELS
# ─────────────────────────────────────────────

class User(UserMixin, db.Model):
    __tablename__ = 'users'
    id            = db.Column(db.Integer, primary_key=True)
    username      = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    role          = db.Column(db.String(20), nullable=False, default='staff')
    full_name     = db.Column(db.String(120))
    created_at    = db.Column(db.DateTime, default=datetime.utcnow)
    is_active     = db.Column(db.Boolean, default=True)

    def set_password(self, pw):
        self.password_hash = generate_password_hash(pw)

    def check_password(self, pw):
        return check_password_hash(self.password_hash, pw)

    @property
    def is_admin(self):
        return self.role == 'superadmin'


class School(db.Model):
    __tablename__ = 'schools'
    id         = db.Column(db.Integer, primary_key=True)
    unique_id  = db.Column(db.String(20), unique=True, nullable=False)
    name       = db.Column(db.String(200), nullable=False)
    location   = db.Column(db.String(200))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    members    = db.relationship('SchoolMember', backref='school',
                                  lazy=True, cascade='all, delete-orphan')


class SchoolMember(db.Model):
    __tablename__ = 'school_members'
    id                = db.Column(db.Integer, primary_key=True)
    unique_id         = db.Column(db.String(20), unique=True, nullable=False)
    school_id         = db.Column(db.Integer, db.ForeignKey('schools.id'), nullable=False)
    full_name         = db.Column(db.String(200), nullable=False)
    role              = db.Column(db.String(50), nullable=False)
    custom_role       = db.Column(db.String(100))
    date_of_birth     = db.Column(db.Date)
    area_of_residence = db.Column(db.String(200))
    phone             = db.Column(db.String(30))
    email             = db.Column(db.String(120))
    photo_path        = db.Column(db.String(300))
    qr_path           = db.Column(db.String(300))
    created_at        = db.Column(db.DateTime, default=datetime.utcnow)
    created_by        = db.Column(db.Integer, db.ForeignKey('users.id'))


class Family(db.Model):
    __tablename__ = 'families'
    id                = db.Column(db.Integer, primary_key=True)
    unique_id         = db.Column(db.String(20), unique=True, nullable=False)
    family_name       = db.Column(db.String(200), nullable=False)
    area_of_residence = db.Column(db.String(200))
    created_at        = db.Column(db.DateTime, default=datetime.utcnow)
    created_by        = db.Column(db.Integer, db.ForeignKey('users.id'))
    members           = db.relationship('FamilyMember', backref='family',
                                         lazy=True, cascade='all, delete-orphan')


class FamilyMember(db.Model):
    __tablename__ = 'family_members'
    id                = db.Column(db.Integer, primary_key=True)
    unique_id         = db.Column(db.String(20), unique=True, nullable=False)
    family_id         = db.Column(db.Integer, db.ForeignKey('families.id'), nullable=False)
    full_name         = db.Column(db.String(200), nullable=False)
    role              = db.Column(db.String(50), nullable=False)
    date_of_birth     = db.Column(db.Date)
    area_of_residence = db.Column(db.String(200))
    phone             = db.Column(db.String(30))
    email             = db.Column(db.String(120))
    photo_path        = db.Column(db.String(300))
    qr_path           = db.Column(db.String(300))
    created_at        = db.Column(db.DateTime, default=datetime.utcnow)
    created_by        = db.Column(db.Integer, db.ForeignKey('users.id'))


class Organization(db.Model):
    __tablename__ = 'organizations'
    id                = db.Column(db.Integer, primary_key=True)
    unique_id         = db.Column(db.String(20), unique=True, nullable=False)
    name              = db.Column(db.String(200), nullable=False)
    area_of_residence = db.Column(db.String(200))
    created_at        = db.Column(db.DateTime, default=datetime.utcnow)
    created_by        = db.Column(db.Integer, db.ForeignKey('users.id'))
    members           = db.relationship('OrgMember', backref='organization',
                                         lazy=True, cascade='all, delete-orphan')


class OrgMember(db.Model):
    __tablename__ = 'org_members'
    id                = db.Column(db.Integer, primary_key=True)
    unique_id         = db.Column(db.String(20), unique=True, nullable=False)
    org_id            = db.Column(db.Integer, db.ForeignKey('organizations.id'), nullable=False)
    full_name         = db.Column(db.String(200), nullable=False)
    role              = db.Column(db.String(100))
    area_of_residence = db.Column(db.String(200))
    phone             = db.Column(db.String(30))
    email             = db.Column(db.String(120))
    photo_path        = db.Column(db.String(300))
    qr_path           = db.Column(db.String(300))
    created_at        = db.Column(db.DateTime, default=datetime.utcnow)
    created_by        = db.Column(db.Integer, db.ForeignKey('users.id'))


class Individual(db.Model):
    __tablename__ = 'individuals'
    id                = db.Column(db.Integer, primary_key=True)
    unique_id         = db.Column(db.String(20), unique=True, nullable=False)
    full_name         = db.Column(db.String(200), nullable=False)
    occupation        = db.Column(db.String(200))
    date_of_birth     = db.Column(db.Date)
    area_of_residence = db.Column(db.String(200))
    phone             = db.Column(db.String(30))
    email             = db.Column(db.String(120))
    photo_path        = db.Column(db.String(300))
    qr_path           = db.Column(db.String(300))
    created_at        = db.Column(db.DateTime, default=datetime.utcnow)
    created_by        = db.Column(db.Integer, db.ForeignKey('users.id'))


class SiteSettings(db.Model):
    __tablename__ = 'site_settings'
    id         = db.Column(db.Integer, primary_key=True)
    key        = db.Column(db.String(100), unique=True, nullable=False)
    value      = db.Column(db.Text)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow,
                            onupdate=datetime.utcnow)


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

@login_manager.user_loader
def load_user(uid):
    return db.session.get(User, int(uid))


def generate_uid(prefix='SL'):
    return f"{prefix}-{uuid.uuid4().hex[:8].upper()}"


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            abort(403)
        return f(*args, **kwargs)
    return decorated


def generate_password(length=14):
    alphabet = string.ascii_letters + string.digits + '!@#$%^'
    while True:
        pwd = ''.join(secrets.choice(alphabet) for _ in range(length))
        if (any(c.islower() for c in pwd)
                and any(c.isupper() for c in pwd)
                and any(c.isdigit() for c in pwd)):
            return pwd


def get_setting(key, default=''):
    s = SiteSettings.query.filter_by(key=key).first()
    return s.value if s else default


def set_setting(key, value):
    s = SiteSettings.query.filter_by(key=key).first()
    if s:
        s.value = value
        s.updated_at = datetime.utcnow()
    else:
        s = SiteSettings(key=key, value=value)
        db.session.add(s)
    db.session.commit()


def generate_qr_code(data: str, uid: str) -> str:
    qr = qrcode.QRCode(version=1, box_size=8, border=2,
                       error_correction=qrcode.constants.ERROR_CORRECT_H)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    fname = f"qr_{uid}.png"
    img.save(os.path.join(app.config['UPLOAD_FOLDER'], fname))
    return fname


def remove_qr(qr_path):
    if qr_path:
        full = os.path.join(app.config['UPLOAD_FOLDER'], qr_path)
        if os.path.exists(full):
            try:
                os.remove(full)
            except OSError:
                pass


def allowed_file(filename):
    return ('.' in filename
            and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS)


def parse_date(s):
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(str(s).strip(), fmt).date()
        except ValueError:
            pass
    return None


def safe_str(val):
    return '' if val is None else str(val).strip()


# ─────────────────────────────────────────────
# EXCEL READING — openpyxl only (no pandas)
# ─────────────────────────────────────────────

def read_excel_sheets(fpath):
    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        all_rows = [[c.value if c.value is not None else ''
                     for c in row] for row in ws.iter_rows()]
        if not all_rows:
            sheets[name] = {'columns': [], 'rows': [], 'total': 0}
            continue
        headers = [str(h) for h in all_rows[0]]
        data    = all_rows[1:]
        sheets[name] = {'columns': headers, 'rows': data, 'total': len(data)}
    wb.close()
    return sheets


def sheet_rows_as_dicts(sheet_data):
    cols = sheet_data['columns']
    return [dict(zip(cols, row)) for row in sheet_data['rows']]


# ─────────────────────────────────────────────
# DASHBOARD STATS
# ─────────────────────────────────────────────

def get_dashboard_stats():
    total_schools  = School.query.count()
    total_students = SchoolMember.query.filter_by(role='Student').count()
    total_families = Family.query.count()
    total_orgs     = Organization.query.count()
    total_indiv    = Individual.query.count()
    total_people   = (SchoolMember.query.count() + FamilyMember.query.count()
                      + OrgMember.query.count() + Individual.query.count())

    category_data = {
        'Schools':       SchoolMember.query.count(),
        'Families':      FamilyMember.query.count(),
        'Organizations': OrgMember.query.count(),
        'Individuals':   Individual.query.count(),
    }

    thirty_ago = datetime.utcnow() - timedelta(days=30)
    daily_regs = []
    for i in range(30):
        day      = thirty_ago + timedelta(days=i)
        next_day = day + timedelta(days=1)
        count = (
            SchoolMember.query.filter(SchoolMember.created_at.between(day, next_day)).count()
            + FamilyMember.query.filter(FamilyMember.created_at.between(day, next_day)).count()
            + OrgMember.query.filter(OrgMember.created_at.between(day, next_day)).count()
            + Individual.query.filter(Individual.created_at.between(day, next_day)).count()
        )
        daily_regs.append({'date': day.strftime('%b %d'), 'count': count})

    top_schools = (
        db.session.query(School.name, func.count(SchoolMember.id).label('cnt'))
        .join(SchoolMember, School.id == SchoolMember.school_id)
        .group_by(School.id)
        .order_by(func.count(SchoolMember.id).desc())
        .limit(5).all()
    )

    recent = []
    for sm in SchoolMember.query.order_by(SchoolMember.created_at.desc()).limit(4).all():
        recent.append({'name': sm.full_name, 'category': 'School', 'role': sm.role,
                       'area': sm.area_of_residence or '',
                       'date': sm.created_at.strftime('%Y-%m-%d'), 'uid': sm.unique_id})
    for fm in FamilyMember.query.order_by(FamilyMember.created_at.desc()).limit(4).all():
        recent.append({'name': fm.full_name, 'category': 'Family', 'role': fm.role,
                       'area': fm.area_of_residence or '',
                       'date': fm.created_at.strftime('%Y-%m-%d'), 'uid': fm.unique_id})
    for om in OrgMember.query.order_by(OrgMember.created_at.desc()).limit(2).all():
        recent.append({'name': om.full_name, 'category': 'Organization',
                       'role': om.role or '', 'area': om.area_of_residence or '',
                       'date': om.created_at.strftime('%Y-%m-%d'), 'uid': om.unique_id})
    for ind in Individual.query.order_by(Individual.created_at.desc()).limit(2).all():
        recent.append({'name': ind.full_name, 'category': 'Individual',
                       'role': ind.occupation or '', 'area': ind.area_of_residence or '',
                       'date': ind.created_at.strftime('%Y-%m-%d'), 'uid': ind.unique_id})
    recent.sort(key=lambda x: x['date'], reverse=True)

    return {
        'total_schools':  total_schools,
        'total_students': total_students,
        'total_families': total_families,
        'total_orgs':     total_orgs,
        'total_indiv':    total_indiv,
        'total_people':   total_people,
        'category_data':  category_data,
        'daily_regs':     daily_regs,
        'top_schools':    [(s[0], s[1]) for s in top_schools],
        'recent':         recent[:10],
    }


# ─────────────────────────────────────────────
# AUTH ROUTES
# ─────────────────────────────────────────────

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = User.query.filter_by(username=username, is_active=True).first()
        if user and user.check_password(password):
            login_user(user, remember=True)
            return redirect(request.args.get('next') or url_for('dashboard'))
        flash('Invalid username or password.', 'error')
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been signed out.', 'info')
    return redirect(url_for('login'))


# ─────────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────────

@app.route('/dashboard')
@login_required
def dashboard():
    stats = get_dashboard_stats()
    return render_template('dashboard.html', stats=stats,
                           company=get_setting('company_name', 'STOPLIGHT SYSTEM'),
                           page='dashboard')


# ─────────────────────────────────────────────
# SCHOOL REGISTRATION
# ─────────────────────────────────────────────

@app.route('/register/school', methods=['GET', 'POST'])
@login_required
def register_school():
    schools = School.query.order_by(School.name).all()
    if request.method == 'POST':
        action = request.form.get('action', 'new')
        if action == 'existing':
            school = db.session.get(School, request.form.get('school_id'))
            if not school:
                flash('School not found.', 'error')
                return redirect(url_for('register_school'))
        else:
            sname = request.form.get('school_name', '').strip()
            sloc  = request.form.get('school_location', '').strip()
            if not sname:
                flash('School name is required.', 'error')
                return redirect(url_for('register_school'))
            if School.query.filter(func.lower(School.name) == sname.lower()).first():
                flash(f'School "{sname}" already exists. Use "Select Existing School".', 'warning')
                return redirect(url_for('register_school'))
            school = School(unique_id=generate_uid('SCH'), name=sname,
                            location=sloc, created_by=current_user.id)
            db.session.add(school)
            db.session.flush()

        names  = request.form.getlist('member_name[]')
        roles  = request.form.getlist('member_role[]')
        custom = request.form.getlist('member_custom_role[]')
        areas  = request.form.getlist('member_area[]')
        dobs   = request.form.getlist('member_dob[]')
        phones = request.form.getlist('member_phone[]')

        added = 0
        for i, name in enumerate(names):
            name = name.strip()
            if not name:
                continue
            role  = roles[i]  if i < len(roles)  else 'Student'
            crole = custom[i] if i < len(custom) else ''
            area  = areas[i]  if i < len(areas)  else ''
            dob   = parse_date(dobs[i] if i < len(dobs) else '')
            phone = phones[i] if i < len(phones) else ''
            uid     = generate_uid('SM')
            qr_file = generate_qr_code(f"STOPLIGHT|SCHOOL|{uid}|{name}|{role}", uid)
            db.session.add(SchoolMember(
                unique_id=uid, school_id=school.id, full_name=name,
                role=role, custom_role=crole, area_of_residence=area,
                date_of_birth=dob, phone=phone, qr_path=qr_file,
                created_by=current_user.id))
            added += 1

        db.session.commit()
        flash(f'Registered {added} member(s) under "{school.name}".', 'success')
        return redirect(url_for('view_school', school_id=school.id))

    return render_template('register_school.html', schools=schools,
                           company=get_setting('company_name', 'STOPLIGHT SYSTEM'),
                           page='school')


@app.route('/schools')
@login_required
def list_schools():
    q       = request.args.get('q', '')
    query   = School.query
    if q:
        query = query.filter(School.name.ilike(f'%{q}%'))
    schools = query.order_by(School.name).all()
    return render_template('list_schools.html', schools=schools, q=q,
                           company=get_setting('company_name', 'STOPLIGHT SYSTEM'),
                           page='list_school')


@app.route('/schools/<int:school_id>')
@login_required
def view_school(school_id):
    school  = db.session.get(School, school_id) or abort(404)
    members = SchoolMember.query.filter_by(school_id=school_id).all()
    return render_template('view_school.html', school=school, members=members,
                           company=get_setting('company_name', 'STOPLIGHT SYSTEM'),
                           page='school')


# ─────────────────────────────────────────────
# FAMILY REGISTRATION
# ─────────────────────────────────────────────

@app.route('/register/family', methods=['GET', 'POST'])
@login_required
def register_family():
    families = Family.query.order_by(Family.family_name).all()
    if request.method == 'POST':
        action = request.form.get('action', 'new')
        if action == 'existing':
            family = db.session.get(Family, request.form.get('family_id'))
            if not family:
                flash('Family not found.', 'error')
                return redirect(url_for('register_family'))
        else:
            fname_input = request.form.get('family_name', '').strip()
            farea       = request.form.get('family_area', '').strip()
            if not fname_input:
                flash('Family name is required.', 'error')
                return redirect(url_for('register_family'))
            if Family.query.filter(func.lower(Family.family_name) == fname_input.lower()).first():
                flash(f'Family "{fname_input}" already exists.', 'warning')
                return redirect(url_for('register_family'))
            family = Family(unique_id=generate_uid('FAM'), family_name=fname_input,
                            area_of_residence=farea, created_by=current_user.id)
            db.session.add(family)
            db.session.flush()

        names  = request.form.getlist('member_name[]')
        roles  = request.form.getlist('member_role[]')
        areas  = request.form.getlist('member_area[]')
        dobs   = request.form.getlist('member_dob[]')
        phones = request.form.getlist('member_phone[]')

        added = 0
        for i, name in enumerate(names):
            name = name.strip()
            if not name:
                continue
            role  = roles[i]  if i < len(roles)  else 'Other'
            area  = areas[i]  if i < len(areas)  else family.area_of_residence or ''
            dob   = parse_date(dobs[i] if i < len(dobs) else '')
            phone = phones[i] if i < len(phones) else ''
            uid     = generate_uid('FM')
            qr_file = generate_qr_code(f"STOPLIGHT|FAMILY|{uid}|{name}|{role}", uid)
            db.session.add(FamilyMember(
                unique_id=uid, family_id=family.id, full_name=name,
                role=role, area_of_residence=area, date_of_birth=dob,
                phone=phone, qr_path=qr_file, created_by=current_user.id))
            added += 1

        db.session.commit()
        flash(f'Registered {added} member(s) to the {family.family_name} family.', 'success')
        return redirect(url_for('view_family', family_id=family.id))

    return render_template('register_family.html', families=families,
                           company=get_setting('company_name', 'STOPLIGHT SYSTEM'),
                           page='family')


@app.route('/families')
@login_required
def list_families():
    q        = request.args.get('q', '')
    query    = Family.query
    if q:
        query = query.filter(Family.family_name.ilike(f'%{q}%'))
    families = query.order_by(Family.family_name).all()
    return render_template('list_families.html', families=families, q=q,
                           company=get_setting('company_name', 'STOPLIGHT SYSTEM'),
                           page='list_family')


@app.route('/families/<int:family_id>')
@login_required
def view_family(family_id):
    family  = db.session.get(Family, family_id) or abort(404)
    members = FamilyMember.query.filter_by(family_id=family_id).all()
    return render_template('view_family.html', family=family, members=members,
                           company=get_setting('company_name', 'STOPLIGHT SYSTEM'),
                           page='family')


# ─────────────────────────────────────────────
# ORGANIZATION REGISTRATION
# ─────────────────────────────────────────────

@app.route('/register/organization', methods=['GET', 'POST'])
@login_required
def register_organization():
    orgs = Organization.query.order_by(Organization.name).all()
    if request.method == 'POST':
        action = request.form.get('action', 'new')
        if action == 'existing':
            org = db.session.get(Organization, request.form.get('org_id'))
            if not org:
                flash('Organization not found.', 'error')
                return redirect(url_for('register_organization'))
        else:
            org_name = request.form.get('org_name', '').strip()
            org_area = request.form.get('org_area', '').strip()
            if not org_name:
                flash('Organization name is required.', 'error')
                return redirect(url_for('register_organization'))
            if Organization.query.filter(func.lower(Organization.name) == org_name.lower()).first():
                flash(f'Organization "{org_name}" already exists.', 'warning')
                return redirect(url_for('register_organization'))
            org = Organization(unique_id=generate_uid('ORG'), name=org_name,
                               area_of_residence=org_area, created_by=current_user.id)
            db.session.add(org)
            db.session.flush()

        names  = request.form.getlist('member_name[]')
        roles  = request.form.getlist('member_role[]')
        areas  = request.form.getlist('member_area[]')
        phones = request.form.getlist('member_phone[]')

        added = 0
        for i, name in enumerate(names):
            name = name.strip()
            if not name:
                continue
            role  = roles[i]  if i < len(roles)  else ''
            area  = areas[i]  if i < len(areas)  else org.area_of_residence or ''
            phone = phones[i] if i < len(phones) else ''
            uid     = generate_uid('OM')
            qr_file = generate_qr_code(f"STOPLIGHT|ORG|{uid}|{name}|{role}", uid)
            db.session.add(OrgMember(
                unique_id=uid, org_id=org.id, full_name=name,
                role=role, area_of_residence=area, phone=phone,
                qr_path=qr_file, created_by=current_user.id))
            added += 1

        db.session.commit()
        flash(f'Registered {added} member(s) to "{org.name}".', 'success')
        return redirect(url_for('view_organization', org_id=org.id))

    return render_template('register_organization.html', orgs=orgs,
                           company=get_setting('company_name', 'STOPLIGHT SYSTEM'),
                           page='organization')


@app.route('/organizations')
@login_required
def list_organizations():
    q    = request.args.get('q', '')
    query = Organization.query
    if q:
        query = query.filter(Organization.name.ilike(f'%{q}%'))
    orgs = query.order_by(Organization.name).all()
    return render_template('list_organizations.html', orgs=orgs, q=q,
                           company=get_setting('company_name', 'STOPLIGHT SYSTEM'),
                           page='list_org')


@app.route('/organizations/<int:org_id>')
@login_required
def view_organization(org_id):
    org     = db.session.get(Organization, org_id) or abort(404)
    members = OrgMember.query.filter_by(org_id=org_id).all()
    return render_template('view_organization.html', org=org, members=members,
                           company=get_setting('company_name', 'STOPLIGHT SYSTEM'),
                           page='organization')


# ─────────────────────────────────────────────
# INDIVIDUAL REGISTRATION
# ─────────────────────────────────────────────

@app.route('/register/individual', methods=['GET', 'POST'])
@login_required
def register_individual():
    if request.method == 'POST':
        full_name  = request.form.get('full_name', '').strip()
        occupation = request.form.get('occupation', '').strip()
        dob        = parse_date(request.form.get('date_of_birth', ''))
        area       = request.form.get('area_of_residence', '').strip()
        phone      = request.form.get('phone', '').strip()
        email      = request.form.get('email', '').strip()

        if not full_name:
            flash('Full name is required.', 'error')
            return redirect(url_for('register_individual'))

        if Individual.query.filter(func.lower(Individual.full_name) == full_name.lower()).first():
            flash(f'"{full_name}" is already registered.', 'warning')
            return redirect(url_for('register_individual'))

        uid     = generate_uid('IND')
        qr_file = generate_qr_code(
            f"STOPLIGHT|INDIVIDUAL|{uid}|{full_name}|{occupation}", uid)
        db.session.add(Individual(
            unique_id=uid, full_name=full_name, occupation=occupation,
            date_of_birth=dob, area_of_residence=area,
            phone=phone, email=email, qr_path=qr_file,
            created_by=current_user.id))
        db.session.commit()
        flash(f'"{full_name}" registered successfully (ID: {uid}).', 'success')
        return redirect(url_for('list_individuals'))

    return render_template('register_individual.html',
                           company=get_setting('company_name', 'STOPLIGHT SYSTEM'),
                           page='individual')


@app.route('/individuals')
@login_required
def list_individuals():
    q    = request.args.get('q', '')
    query = Individual.query
    if q:
        query = query.filter(or_(
            Individual.full_name.ilike(f'%{q}%'),
            Individual.occupation.ilike(f'%{q}%'),
            Individual.area_of_residence.ilike(f'%{q}%')))
    inds = query.order_by(Individual.full_name).all()
    return render_template('list_individuals.html', individuals=inds, q=q,
                           company=get_setting('company_name', 'STOPLIGHT SYSTEM'),
                           page='list_indiv')


# ─────────────────────────────────────────────
# EXCEL IMPORT  (openpyxl only — no pandas)
# ─────────────────────────────────────────────

@app.route('/import', methods=['GET', 'POST'])
@login_required
def import_excel():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected.', 'error')
            return redirect(url_for('import_excel'))
        f = request.files['file']
        if not f.filename or not allowed_file(f.filename):
            flash('Please upload a valid Excel file (.xlsx or .xls).', 'error')
            return redirect(url_for('import_excel'))

        filename = secure_filename(f.filename)
        fpath    = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        f.save(fpath)

        try:
            sheets  = read_excel_sheets(fpath)
            preview = {
                name: {
                    'columns': d['columns'],
                    'rows':    d['rows'][:5],
                    'total':   d['total'],
                }
                for name, d in sheets.items()
            }
            session['import_file'] = fpath
            return render_template('import_preview.html', preview=preview,
                                   filename=filename,
                                   company=get_setting('company_name', 'STOPLIGHT SYSTEM'),
                                   page='import')
        except Exception as e:
            flash(f'Error reading file: {e}', 'error')
            return redirect(url_for('import_excel'))

    return render_template('import_excel.html',
                           company=get_setting('company_name', 'STOPLIGHT SYSTEM'),
                           page='import')


@app.route('/import/confirm', methods=['POST'])
@login_required
def import_confirm():
    fpath = session.get('import_file')
    if not fpath or not os.path.exists(fpath):
        flash('Import session expired. Please re-upload.', 'error')
        return redirect(url_for('import_excel'))

    try:
        sheets = read_excel_sheets(fpath)
    except Exception as e:
        flash(f'Could not read file: {e}', 'error')
        return redirect(url_for('import_excel'))

    imported, errors = 0, []

    for sheet_name, sheet_data in sheets.items():
        rows = sheet_rows_as_dicts(sheet_data)
        cat  = sheet_name.strip().lower()

        if 'individual' in cat:
            for row in rows:
                try:
                    name = safe_str(row.get('Full Name') or row.get('full_name'))
                    if not name:
                        continue
                    if Individual.query.filter(
                            func.lower(Individual.full_name) == name.lower()).first():
                        continue
                    uid     = generate_uid('IND')
                    qr_file = generate_qr_code(f"STOPLIGHT|INDIVIDUAL|{uid}|{name}", uid)
                    db.session.add(Individual(
                        unique_id=uid, full_name=name,
                        occupation=safe_str(row.get('Occupation')),
                        date_of_birth=parse_date(row.get('Date of Birth')),
                        area_of_residence=safe_str(row.get('Area of Residence')),
                        phone=safe_str(row.get('Phone')),
                        qr_path=qr_file, created_by=current_user.id))
                    imported += 1
                except Exception as e:
                    errors.append(str(e))

        elif 'school' in cat:
            for row in rows:
                try:
                    sname = safe_str(row.get('School Name') or row.get('school_name')) or 'Imported School'
                    mname = safe_str(row.get('Full Name') or row.get('full_name'))
                    if not mname:
                        continue
                    school = School.query.filter(
                        func.lower(School.name) == sname.lower()).first()
                    if not school:
                        school = School(unique_id=generate_uid('SCH'), name=sname,
                                        location=safe_str(row.get('Location')),
                                        created_by=current_user.id)
                        db.session.add(school)
                        db.session.flush()
                    uid     = generate_uid('SM')
                    role    = safe_str(row.get('Role')) or 'Student'
                    qr_file = generate_qr_code(
                        f"STOPLIGHT|SCHOOL|{uid}|{mname}|{role}", uid)
                    db.session.add(SchoolMember(
                        unique_id=uid, school_id=school.id, full_name=mname,
                        role=role,
                        area_of_residence=safe_str(row.get('Area of Residence')),
                        qr_path=qr_file, created_by=current_user.id))
                    imported += 1
                except Exception as e:
                    errors.append(str(e))

        elif 'family' in cat:
            for row in rows:
                try:
                    fname = safe_str(row.get('Family Name') or row.get('family_name')) or 'Imported Family'
                    mname = safe_str(row.get('Full Name') or row.get('full_name'))
                    if not mname:
                        continue
                    fam = Family.query.filter(
                        func.lower(Family.family_name) == fname.lower()).first()
                    if not fam:
                        fam = Family(unique_id=generate_uid('FAM'), family_name=fname,
                                     area_of_residence=safe_str(row.get('Area of Residence')),
                                     created_by=current_user.id)
                        db.session.add(fam)
                        db.session.flush()
                    uid     = generate_uid('FM')
                    role    = safe_str(row.get('Role')) or 'Member'
                    qr_file = generate_qr_code(
                        f"STOPLIGHT|FAMILY|{uid}|{mname}|{role}", uid)
                    db.session.add(FamilyMember(
                        unique_id=uid, family_id=fam.id, full_name=mname,
                        role=role,
                        area_of_residence=safe_str(row.get('Area of Residence')),
                        qr_path=qr_file, created_by=current_user.id))
                    imported += 1
                except Exception as e:
                    errors.append(str(e))

        elif 'org' in cat:
            for row in rows:
                try:
                    oname = safe_str(row.get('Organization Name') or row.get('org_name')) or 'Imported Org'
                    mname = safe_str(row.get('Full Name') or row.get('full_name'))
                    if not mname:
                        continue
                    org = Organization.query.filter(
                        func.lower(Organization.name) == oname.lower()).first()
                    if not org:
                        org = Organization(unique_id=generate_uid('ORG'), name=oname,
                                           area_of_residence=safe_str(row.get('Area of Residence')),
                                           created_by=current_user.id)
                        db.session.add(org)
                        db.session.flush()
                    uid     = generate_uid('OM')
                    role    = safe_str(row.get('Role'))
                    qr_file = generate_qr_code(
                        f"STOPLIGHT|ORG|{uid}|{mname}|{role}", uid)
                    db.session.add(OrgMember(
                        unique_id=uid, org_id=org.id, full_name=mname,
                        role=role,
                        area_of_residence=safe_str(row.get('Area of Residence')),
                        phone=safe_str(row.get('Phone')),
                        qr_path=qr_file, created_by=current_user.id))
                    imported += 1
                except Exception as e:
                    errors.append(str(e))

    db.session.commit()
    cat = 'success' if not errors else 'warning'
    flash(f'Import complete: {imported} records added, {len(errors)} error(s).', cat)
    return redirect(url_for('dashboard'))


# ─────────────────────────────────────────────
# ID CARD + QR CODE
# ─────────────────────────────────────────────

def _resolve_record(category, record_id):
    if category == 'school':
        r = db.session.get(SchoolMember, record_id)
        return r, 'School', (r.school.name if r else ''), \
               (r.custom_role or r.role if r else '')
    if category == 'family':
        r = db.session.get(FamilyMember, record_id)
        return r, 'Family', (r.family.family_name if r else ''), \
               (r.role if r else '')
    if category == 'org':
        r = db.session.get(OrgMember, record_id)
        return r, 'Organization', (r.organization.name if r else ''), \
               (r.role or '' if r else '')
    if category == 'individual':
        r = db.session.get(Individual, record_id)
        return r, 'Individual', '', (r.occupation or '' if r else '')
    return None, '', '', ''


def _b64_file(path):
    if path and os.path.exists(path):
        with open(path, 'rb') as fh:
            return base64.b64encode(fh.read()).decode()
    return ''


@app.route('/id-card/<category>/<int:record_id>')
@login_required
def view_id_card(category, record_id):
    record, cat_name, group, _ = _resolve_record(category, record_id)
    if not record:
        abort(404)
    qr_b64   = _b64_file(os.path.join(app.config['UPLOAD_FOLDER'],
                                       record.qr_path or ''))
    logo_b64 = _b64_file(get_setting('logo_path', ''))
    return render_template('id_card.html', record=record, category=category,
                           cat_name=cat_name, group=group,
                           qr_img_b64=qr_b64, logo_b64=logo_b64,
                           company=get_setting('company_name', 'STOPLIGHT SYSTEM'),
                           page='id_card')


@app.route('/id-card/pdf/<category>/<int:record_id>')
@login_required
def download_id_card_pdf(category, record_id):
    record, cat_name, group, role = _resolve_record(category, record_id)
    if not record:
        abort(404)

    buf    = io.BytesIO()
    card_w = 86 * mm
    card_h = 54 * mm
    doc    = SimpleDocTemplate(buf, pagesize=(card_w, card_h),
                               leftMargin=4*mm, rightMargin=4*mm,
                               topMargin=4*mm, bottomMargin=4*mm)
    company  = get_setting('company_name', 'STOPLIGHT SYSTEM')
    title_st = ParagraphStyle('t',  fontSize=7, fontName='Helvetica-Bold',
                               alignment=TA_CENTER,
                               textColor=colors.HexColor('#0a2342'), spaceAfter=1*mm)
    name_st  = ParagraphStyle('n',  fontSize=9, fontName='Helvetica-Bold',
                               alignment=TA_CENTER,
                               textColor=colors.HexColor('#0a2342'), spaceAfter=1*mm)
    body_st  = ParagraphStyle('b',  fontSize=6.5, fontName='Helvetica',
                               alignment=TA_CENTER,
                               textColor=colors.HexColor('#333333'), spaceAfter=0.5*mm)
    id_st    = ParagraphStyle('id', fontSize=6, fontName='Helvetica-Bold',
                               alignment=TA_CENTER,
                               textColor=colors.HexColor('#1b6ca8'))
    story = []
    logo_path = get_setting('logo_path', '')
    if logo_path and os.path.exists(logo_path):
        story.append(RLImage(logo_path, width=12*mm, height=8*mm))
    story.append(Paragraph(company.upper(), title_st))
    story.append(HRFlowable(width='100%', thickness=0.5,
                             color=colors.HexColor('#1b6ca8')))
    story.append(Spacer(1, 1*mm))
    if record.qr_path:
        qr_full = os.path.join(app.config['UPLOAD_FOLDER'], record.qr_path)
        if os.path.exists(qr_full):
            story.append(RLImage(qr_full, width=18*mm, height=18*mm))
    story += [
        Spacer(1, 1*mm),
        Paragraph(record.full_name.upper(), name_st),
        Paragraph(role.upper(), body_st),
        Paragraph(f"{cat_name}{(' | ' + group) if group else ''}", body_st),
        Spacer(1, 1*mm),
        HRFlowable(width='100%', thickness=0.3, color=colors.grey),
        Paragraph(f"ID: {record.unique_id}", id_st),
        Paragraph('Signature: _________________', body_st),
    ]
    doc.build(story)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"ID_{record.unique_id}.pdf",
                     mimetype='application/pdf')


# ─────────────────────────────────────────────
# PRINT PAGES
# ─────────────────────────────────────────────

@app.route('/print/school/<int:school_id>')
@login_required
def print_school(school_id):
    school  = db.session.get(School, school_id) or abort(404)
    members = SchoolMember.query.filter_by(school_id=school_id).all()
    return render_template('print_school.html', school=school, members=members,
                           company=get_setting('company_name', 'STOPLIGHT SYSTEM'),
                           logo_path=get_setting('logo_path', ''))


@app.route('/print/family/<int:family_id>')
@login_required
def print_family(family_id):
    family  = db.session.get(Family, family_id) or abort(404)
    members = FamilyMember.query.filter_by(family_id=family_id).all()
    return render_template('print_family.html', family=family, members=members,
                           company=get_setting('company_name', 'STOPLIGHT SYSTEM'),
                           logo_path=get_setting('logo_path', ''))


@app.route('/print/individuals')
@login_required
def print_individuals():
    individuals = Individual.query.order_by(Individual.full_name).all()
    return render_template('print_individuals.html', individuals=individuals,
                           company=get_setting('company_name', 'STOPLIGHT SYSTEM'),
                           logo_path=get_setting('logo_path', ''))


# ─────────────────────────────────────────────
# SETTINGS  (Super Admin only)
# ─────────────────────────────────────────────

@app.route('/settings', methods=['GET', 'POST'])
@login_required
@admin_required
def settings():
    if request.method == 'POST':
        cname = request.form.get('company_name', '').strip()
        if cname:
            set_setting('company_name', cname)
        logo = request.files.get('logo')
        if logo and logo.filename:
            ext = logo.filename.rsplit('.', 1)[-1].lower()
            if ext in {'png', 'jpg', 'jpeg'}:
                fname = f"logo.{ext}"
                lpath = os.path.join(app.config['UPLOAD_FOLDER'], fname)
                logo.save(lpath)
                set_setting('logo_path', lpath)
        flash('Settings saved.', 'success')
        return redirect(url_for('settings'))

    users = User.query.order_by(User.username).all()
    return render_template('settings.html',
                           company=get_setting('company_name', 'STOPLIGHT SYSTEM'),
                           logo_path=get_setting('logo_path', ''),
                           users=users, page='settings')


@app.route('/settings/add-user', methods=['POST'])
@login_required
@admin_required
def add_user():
    username  = request.form.get('username', '').strip()
    full_name = request.form.get('full_name', '').strip()
    role      = request.form.get('role', 'staff')
    password  = request.form.get('password', '').strip() or generate_password()

    if not username:
        flash('Username is required.', 'error')
        return redirect(url_for('settings'))
    if User.query.filter_by(username=username).first():
        flash(f'Username "{username}" already exists.', 'error')
        return redirect(url_for('settings'))

    user = User(username=username, full_name=full_name, role=role)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    flash(f'User "{username}" created. Password: {password}', 'success')
    return redirect(url_for('settings'))


@app.route('/settings/toggle-user/<int:uid>', methods=['POST'])
@login_required
@admin_required
def toggle_user(uid):
    user = db.session.get(User, uid)
    if user and user.id != current_user.id:
        user.is_active = not user.is_active
        db.session.commit()
        state = 'activated' if user.is_active else 'deactivated'
        flash(f'User "{user.username}" {state}.', 'success')
    return redirect(url_for('settings'))


# ─────────────────────────────────────────────
# DELETE ROUTES
# ─────────────────────────────────────────────

@app.route('/delete/school-member/<int:member_id>', methods=['POST'])
@login_required
def delete_school_member(member_id):
    member = db.session.get(SchoolMember, member_id) or abort(404)
    school_id = member.school_id
    remove_qr(member.qr_path)
    name = member.full_name
    db.session.delete(member)
    db.session.commit()
    flash(f'Member "{name}" deleted.', 'success')
    return redirect(url_for('view_school', school_id=school_id))


@app.route('/delete/school/<int:school_id>', methods=['POST'])
@login_required
@admin_required
def delete_school(school_id):
    school = db.session.get(School, school_id) or abort(404)
    for m in school.members:
        remove_qr(m.qr_path)
    name = school.name
    db.session.delete(school)
    db.session.commit()
    flash(f'School "{name}" and all members permanently deleted.', 'success')
    return redirect(url_for('list_schools'))


@app.route('/delete/family-member/<int:member_id>', methods=['POST'])
@login_required
def delete_family_member(member_id):
    member = db.session.get(FamilyMember, member_id) or abort(404)
    family_id = member.family_id
    remove_qr(member.qr_path)
    name = member.full_name
    db.session.delete(member)
    db.session.commit()
    flash(f'Family member "{name}" deleted.', 'success')
    return redirect(url_for('view_family', family_id=family_id))


@app.route('/delete/family/<int:family_id>', methods=['POST'])
@login_required
@admin_required
def delete_family(family_id):
    family = db.session.get(Family, family_id) or abort(404)
    for m in family.members:
        remove_qr(m.qr_path)
    name = family.family_name
    db.session.delete(family)
    db.session.commit()
    flash(f'Family "{name}" and all members permanently deleted.', 'success')
    return redirect(url_for('list_families'))


@app.route('/delete/org-member/<int:member_id>', methods=['POST'])
@login_required
def delete_org_member(member_id):
    member = db.session.get(OrgMember, member_id) or abort(404)
    org_id = member.org_id
    remove_qr(member.qr_path)
    name = member.full_name
    db.session.delete(member)
    db.session.commit()
    flash(f'Member "{name}" deleted.', 'success')
    return redirect(url_for('view_organization', org_id=org_id))


@app.route('/delete/organization/<int:org_id>', methods=['POST'])
@login_required
@admin_required
def delete_organization(org_id):
    org = db.session.get(Organization, org_id) or abort(404)
    for m in org.members:
        remove_qr(m.qr_path)
    name = org.name
    db.session.delete(org)
    db.session.commit()
    flash(f'Organization "{name}" and all members permanently deleted.', 'success')
    return redirect(url_for('list_organizations'))


@app.route('/delete/individual/<int:ind_id>', methods=['POST'])
@login_required
def delete_individual(ind_id):
    ind = db.session.get(Individual, ind_id) or abort(404)
    remove_qr(ind.qr_path)
    name = ind.full_name
    db.session.delete(ind)
    db.session.commit()
    flash(f'Individual "{name}" deleted.', 'success')
    return redirect(url_for('list_individuals'))


# ─────────────────────────────────────────────
# JSON / AUTO-SUGGEST API
# ─────────────────────────────────────────────

@app.route('/api/schools')
@login_required
def api_schools():
    q    = request.args.get('q', '')
    rows = School.query.filter(School.name.ilike(f'%{q}%')).limit(10).all()
    return jsonify([{'id': s.id, 'name': s.name, 'location': s.location} for s in rows])


@app.route('/api/families')
@login_required
def api_families():
    q    = request.args.get('q', '')
    rows = Family.query.filter(Family.family_name.ilike(f'%{q}%')).limit(10).all()
    return jsonify([{'id': f.id, 'name': f.family_name} for f in rows])


@app.route('/api/orgs')
@login_required
def api_orgs():
    q    = request.args.get('q', '')
    rows = Organization.query.filter(Organization.name.ilike(f'%{q}%')).limit(10).all()
    return jsonify([{'id': o.id, 'name': o.name} for o in rows])


@app.route('/api/check-duplicate')
@login_required
def api_check_duplicate():
    name  = request.args.get('name', '').strip()
    cat   = request.args.get('category', 'individual')
    found = False
    if cat == 'individual':
        found = bool(Individual.query.filter(
            func.lower(Individual.full_name) == name.lower()).first())
    elif cat == 'school_member':
        found = bool(SchoolMember.query.filter(
            func.lower(SchoolMember.full_name) == name.lower()).first())
    elif cat == 'family_member':
        found = bool(FamilyMember.query.filter(
            func.lower(FamilyMember.full_name) == name.lower()).first())
    return jsonify({'duplicate': found})


# ─────────────────────────────────────────────
# DATABASE INIT + SEED
# ─────────────────────────────────────────────

def init_db():
    with app.app_context():
        db.create_all()

        if not User.query.filter_by(role='superadmin').first():
            admin_pass = generate_password(16)
            staff_pass = generate_password(14)

            admin = User(username='superadmin',
                         full_name='System Administrator', role='superadmin')
            admin.set_password(admin_pass)

            staff = User(username='staffuser',
                         full_name='Default Staff', role='staff')
            staff.set_password(staff_pass)

            db.session.add_all([admin, staff])
            db.session.commit()

            creds = (
                "=" * 50 + "\n"
                "STOPLIGHT SYSTEM — INITIAL CREDENTIALS\n"
                "=" * 50 + "\n\n"
                f"SUPER ADMIN\n  Username: superadmin\n  Password: {admin_pass}\n\n"
                f"STAFF USER\n  Username: staffuser\n  Password: {staff_pass}\n\n"
                "IMPORTANT: Change these passwords after first login.\n"
                "=" * 50 + "\n"
            )
            cred_path = os.path.join(BASE_DIR, 'CREDENTIALS.txt')
            with open(cred_path, 'w') as cf:
                cf.write(creds)
            print(creds)

        if not SiteSettings.query.filter_by(key='company_name').first():
            db.session.add(SiteSettings(key='company_name', value='STOPLIGHT SYSTEM'))
            db.session.commit()


if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
